# -*- coding: utf-8 -*-

import sys
import requests
import linecache
from bs4 import BeautifulSoup
from openpyxl import Workbook,load_workbook
from tqdm import tqdm
import time
import random
from multiprocessing import Process,Pool,Manager
from xlutils.copy import copy
import os
reload(sys)
sys.setdefaultencoding('utf-8')
# 获取文件制定行数，是第一次读取后面直接从缓存中拿
def get_line(filePath, line):
  data = linecache.getline(filePath, line)
  line_date = data.split(',')
  return line_date

def request_line(startInfo = ['','','', ''], endInfo = ['','','', ''], composeData = []):
  param = {
    '_fm.es._0.s': startInfo[0],
    '_fm.es._0.se': startInfo[1],
    '_fm.es._0.sen': startInfo[2],
    '_fm.es._0.r': endInfo[0],
    '_fm.es._0.re': endInfo[1],
    '_fm.es._0.rec': endInfo[2]
  }
  try:
    proxy = get_random_ip(proxy_list)
    r = requests.get('https://56.1688.com/order/price/estimate_price.htm?notFirst=true&_fm.es._0.o=&_fm.es._0.c=&_fm.es._0.co=&_fm.es._0.sp=&_fm.es._0.i=&_fm.es._0.is=&_fm.es._0.isf=&_fm.es._0.isp=&_fm.es._0.iscn=&_fm.es._0.isn=&_fm.es._0.isc=&_fm.es._0.isa=&_fm.es._0.ro=&_fm.es._0.w=&_fm.es._0.we=&_fm.es._0.v=&_fm.es._0.vo=&_fm.es._0.tra=&_fm.es._0.tran=&_fm.es._0.isfr=&_fm.es._0.a=&_fm.es._0.d=&_fm.es._0.d=&_fm.es._0.d=&_fm.es._0.d=&r=1560672972573&sizePerPage=&pageIndex=1&weight=&expressRouteSortType=&_fm.es._0.sentc=',proxies=proxy ,params=param)
    soup = BeautifulSoup(r.text, 'lxml')
    tags = soup.select('#orderbyBar span[style="font-weight:bold;color:#FF7300"]')
    composeData.append([startInfo[3].replace('\n', '').decode('utf-8'), endInfo[3].replace('\n', '').decode('utf-8'),tags[0].get_text().decode('utf-8'), tags[1].get_text().decode('utf-8')])
  except Exception as e:
    # print(e, r.text)
    print('\n'+startInfo[3]+'-'+endInfo[3]+'\n')
  finally:
    return composeData

# 读取ip文件
def get_ip_list():
  temp_ip_list=[]
  with open('iplist.txt', 'r') as f:
    while True:
      ip=f.readline().replace('\n','')
      temp_ip_list.append('http://'+ip)
      if not ip:
          break
  return temp_ip_list

# 获取随机代理ip
def get_random_ip(list):      
  proxy_ip=random.choice(list)
  proxies={
      'http':proxy_ip
  }
  return proxies

# 组合对应的请求数据参数
def pollData():
  queryInfo = []
  for line in range(start, end+1):
    resData = get_line('/Users/lavi/Documents/cplus/city_province.txt', line)
    for inLine in range(start, end+1):
      toData = get_line('/Users/lavi/Documents/cplus/city_province.txt', inLine)
      queryInfo.append([[resData[0], resData[2],'', resData[1]+resData[3]], [toData[0], toData[2],'', toData[1]+toData[3]]])
  return queryInfo

# 初始化对应的数据
def init():
  global start
  global end
  global pbar
  global sheet_name
  global slice_num # 数组切割次数
  global proxy_list
  slice_num = 300
  sheet_name = '0712.xlsx'
  wb = Workbook()
  wb.active.append(['起始位置','终点位置', '公司数量', '线路数量'])
  wb.save(sheet_name)
  start = input('请输入开始行数:')
  end = input('请输入结束行数:')
  proxy_list = get_ip_list()
  pbar = tqdm(total=((end-start+1) * (end-start+1)))

#更新进度条 并且进行追加数据到excel中
def updatePb(res):
  pbar.update(1)
  if (len(res) >= slice_num):
    saveFile_append(res)
# 追加写入excel
def saveFile_append(res):
  workbook = load_workbook(sheet_name)  # 打开工作簿 保留原有格式
  sheets = workbook.active  # 获取工作簿中的所有表格
  for i in res[0:slice_num+1]:
    sheets.append(i)
  workbook.save(sheet_name)  # 保存工作簿
  removeList(res)

# 移除数组中的已经写入的数据
def removeList(composeData):
  for i in composeData[0:(slice_num+1)]:
    composeData.remove(i)
def main():
  init()
  global composeData
  compose = pollData()
  manager = Manager()
  composeData = manager.list([]) # 进程数据共享
  pool = Pool(4)
  print '\nParent process %s'%os.getpid()
  for i in compose:
    pool.apply_async(request_line, args=(i[0],i[1],composeData),callback=updatePb)
  pool.close()
  pool.join()
  pbar.close()
  saveFile_append(composeData) # 最后将剩余的数据写入excel中
  print '\ndone'

if __name__=='__main__':
  main()